import openpyxl
import os
import time
from datetime import date,datetime

def open_sheet():
	global wb,folderPath,sheet
	m=["Beacons"]
	folderName = "MOTES_DATABASE"
	folderPath = os.path.join(os.path.dirname(os.path.realpath('__file__')), folderName)
	wb=openpyxl.Workbook()
	 
	sheet = wb.active
	if not os.path.exists(folderPath):
		  os.makedirs(folderPath)
		  
	try:
		wb = openpyxl.load_workbook(folderPath+"/"+"BEACONS.xlsx")
	except:
		 wb.save(folderPath+"/"+"BEACONS.xlsx") 
		 x = sheet.cell(row = 1, column = 1)
		 x.value =m[0]

def excel_put(l):
 open_sheet()
 sheet = wb.active
 s=sheet.max_row+1
 print(s)

 y = sheet.cell(row =s, column = 1)
 y.value=l

 print("Beacon added in Database")

 
 
 wb.save(folderPath+"/"+"BEACONS.xlsx")
 return   


def excel_check(value):
 open_sheet()
 sheet = wb.active
 s=sheet.max_row

 print(s)

 if value in [sheet.cell(row =i, column = 1).value for i in range(1,s+1)]:
  print("Key match found---------")
  wb.save(folderPath+"/"+"BEACONS.xlsx")
  return "found"   
 else:
  print("Not found---")  
  wb.save(folderPath+"/"+"BEACONS.xlsx")
  return "not found" 

 
